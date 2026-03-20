"""Finance module web routes."""
from __future__ import annotations

import asyncio
import csv
import io
import logging
from datetime import datetime, timezone, timedelta
from decimal import Decimal
from typing import AsyncGenerator

from jose import jwt as _jwt_lib
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import JSONResponse, StreamingResponse
from itsdangerous import BadSignature, SignatureExpired, TimestampSigner
from sqlalchemy.ext.asyncio import AsyncSession

from shared.config import settings
from shared.db import get_async_session

logger = logging.getLogger(__name__)
from shared.services.pin_guard import record_pin_fail, clear_pin_fail, should_alert as _should_alert
from shared.services.finance_pin import (
    get_finance_settings as _get_finance_settings_row,
    verify_finance_pin,
    set_finance_pin,
    reset_finance_pin,
    get_cash_balance as _get_cash_balance,
    set_cash_balance as _set_cash_balance,
)
from shared.services.finance_service import (
    list_categories as _list_cats,
    create_category as _create_cat,
    update_category as _update_cat,
    list_operations as _list_ops,
    get_operation as _get_op,
    create_operation as _create_op,
    update_operation as _update_op,
    delete_operation as _delete_op,
    get_dashboard as _get_dash,
    get_avg_expense_last_7_days as _get_avg_exp7,
    export_operations as _export_ops,
    _serialize_operation,
)
from shared.services.warehouse import get_warehouse_value_rub as _get_warehouse_value_rub
from sqlalchemy import select, func as _func
from shared.models import FinanceOperation as _FinOp, SalaryPayout as _SalaryPayout
from shared.utils import utc_now

router = APIRouter()

def _delta_pct(current, prev) -> float | None:
    """Return % change (current vs prev), or None when prev==0 and current!=0."""
    try:
        c = float(current)
        p = float(prev)
        if abs(p) < 0.001:
            return None if abs(c) > 0.001 else 0.0
        return round((c - p) / abs(p) * 100, 1)
    except Exception:
        return None


# ── PIN helpers ───────────────────────────────────────────────────────────────

_COOKIE = "finance_pin_ok"
_TTL = 72 * 3600


def _human_error(code: str) -> str:
    c = str(code or "").strip()
    mapping = {
        "wrong_pin": "Неверный PIN-код.",
        "invalid_pin": "PIN должен состоять из 6 цифр.",
        "invalid_amount": "Введите корректную сумму больше нуля.",
        "invalid_type": "Некорректный тип операции.",
        "not_found": "Запись не найдена.",
        "name_required": "Введите название.",
    }
    return mapping.get(c, "Произошла ошибка. Попробуйте ещё раз.")


def _signer() -> TimestampSigner:
    return TimestampSigner(str(getattr(settings, "WEB_JWT_SECRET", "") or "") + "_finance")


def pin_valid(request: Request) -> bool:
    token = str(request.cookies.get(_COOKIE) or "").strip()
    if not token:
        return False
    try:
        val = _signer().unsign(token, max_age=_TTL)
        return str(val.decode()).strip() == "1"
    except (BadSignature, SignatureExpired):
        return False
    except Exception:
        return False


def _set_pin_cookie(resp: JSONResponse) -> None:
    token = _signer().sign("1").decode()
    resp.set_cookie(_COOKIE, token, max_age=_TTL, httponly=True, secure=False, samesite="lax", path="/")


def _get_user_key(request: Request) -> tuple[str, str]:
    """Returns (cache_key, display_name) for the requesting admin."""
    token = request.cookies.get("admin_token")
    if token:
        try:
            data = _jwt_lib.decode(token, settings.WEB_JWT_SECRET, algorithms=["HS256"])
            uid = str(data.get("sub") or "")
            if uid:
                return f"uid:{uid}", f"user_id={uid}"
        except Exception:
            pass
    ip = str(request.client.host) if request.client else "unknown"
    return f"ip:{ip}", f"ip={ip}"


async def _send_pin_alert(user_display: str, attempts: int) -> None:
    from web.app.services.messenger import Messenger
    from datetime import datetime as _dt
    token = str(getattr(settings, "BOT_TOKEN", "") or "").strip()
    if not token:
        return
    admin_ids = list(getattr(settings, "admin_ids", None) or [])
    if not admin_ids:
        return
    now_str = _dt.now().strftime("%d.%m.%Y %H:%M:%S")
    text = (
        f"🔐 <b>Финансы: подозрительный ввод PIN</b>\n"
        f"👤 Кто: <code>{user_display}</code>\n"
        f"❌ Неверных попыток: {attempts}\n"
        f"🕐 Время: {now_str}\n"
        f"🔒 Раздел: Финансы"
    )
    messenger = Messenger(token)
    n_sent = 0
    for uid in admin_ids:
        try:
            await messenger.send_message(chat_id=int(uid), text=text, parse_mode="HTML")
            n_sent += 1
        except Exception:
            pass
    logger.info("FINANCE_PIN_ALERT_SENT admins=%d", n_sent)


# ── Lazy deps ─────────────────────────────────────────────────────────────────

async def get_db() -> AsyncGenerator[AsyncSession, None]:
    async with get_async_session() as session:
        yield session


async def _require_pin(request: Request) -> None:
    if not pin_valid(request):
        raise HTTPException(status_code=403)


# ── PIN API ───────────────────────────────────────────────────────────────────

@router.post("/api/finance/pin/verify")
async def fin_pin_verify(request: Request, session: AsyncSession = Depends(get_db)):
    user_key, user_display = _get_user_key(request)
    try:
        body = await request.json()
    except Exception:
        body = {}
    ok = await verify_finance_pin(session=session, pin=str(body.get("pin") or "").strip())
    if ok:
        logger.info("FINANCE_PIN_VERIFY ok=True user=%s", user_display)
        clear_pin_fail(user_key)
        resp = JSONResponse({"ok": True})
        _set_pin_cookie(resp)
        logger.info("FINANCE_PIN_SESSION_SET user=%s", user_display)
        return resp
    count = record_pin_fail(user_key)
    logger.info("FINANCE_PIN_FAIL attempt=%d/%d user=%s", count, 3, user_display)
    if _should_alert(count):
        asyncio.create_task(_send_pin_alert(user_display, count))
    return JSONResponse({"ok": False, "error": "wrong_pin", "error_message": _human_error("wrong_pin")}, status_code=403)


@router.post("/api/finance/pin/set")
async def fin_pin_set(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    try:
        body = await request.json()
    except Exception:
        body = {}
    try:
        await set_finance_pin(session=session, new_pin=str(body.get("pin") or "").strip(), updated_by_user_id=None)
    except ValueError:
        return JSONResponse({"ok": False, "error": "invalid_pin", "error_message": _human_error("invalid_pin")}, status_code=400)
    return {"ok": True}


@router.post("/api/finance/pin/reset")
async def fin_pin_reset(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    await reset_finance_pin(session=session, updated_by_user_id=None)
    return {"ok": True}


# ── Dashboard ─────────────────────────────────────────────────────────────────

@router.get("/api/finance/dashboard")
async def fin_dashboard(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    qp = request.query_params
    date_from_s = str(qp.get("from") or "").strip()
    date_to_s = str(qp.get("to") or "").strip()
    now = datetime.now(timezone.utc)
    try:
        date_from = datetime.fromisoformat(date_from_s) if date_from_s else now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    except Exception:
        date_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    try:
        date_to = datetime.fromisoformat(date_to_s) if date_to_s else now
    except Exception:
        date_to = now
    duration = max(1, (date_to.date() - date_from.date()).days + 1)
    prev_from = date_from.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=duration)
    prev_to = date_to.replace(hour=23, minute=59, second=59, microsecond=0) - timedelta(days=duration)

    dash = await _get_dash(session=session, date_from=date_from, date_to=date_to)
    prev = await _get_dash(session=session, date_from=prev_from, date_to=prev_to)
    avg_exp7 = await _get_avg_exp7(session=session)
    cash_bal = await _get_cash_balance(session=session)
    warehouse_val = await _get_warehouse_value_rub(session=session)

    try:
        cushion_days = round(float(cash_bal) / float(avg_exp7), 1) if avg_exp7 > 0 else None
    except Exception:
        cushion_days = None

    prev_exp_map = {r["category_id"]: float(r["total"]) for r in prev.expense_by_category}
    prev_inc_map = {r["category_id"]: float(r["total"]) for r in prev.income_by_category}
    exp_cats = [
        {**r, "prev_total": f"{prev_exp_map.get(r['category_id'], 0):.2f}",
         "delta_pct": _delta_pct(float(r["total"]), prev_exp_map.get(r["category_id"], 0))}
        for r in dash.expense_by_category
    ]
    inc_cats = [
        {**r, "prev_total": f"{prev_inc_map.get(r['category_id'], 0):.2f}",
         "delta_pct": _delta_pct(float(r["total"]), prev_inc_map.get(r["category_id"], 0))}
        for r in dash.income_by_category
    ]

    return {
        "ok": True,
        "period": {"from": date_from.date().isoformat(), "to": date_to.date().isoformat()},
        "prev_period": {"from": prev_from.date().isoformat(), "to": prev_to.date().isoformat()},
        "income": f"{dash.income:.2f}",
        "expense": f"{dash.expense:.2f}",
        "profit": f"{dash.profit:.2f}",
        "avg_expense_per_day": f"{dash.avg_expense_per_day:.2f}",
        "avg_income_per_day": f"{dash.avg_income_per_day:.2f}",
        "cash_balance": f"{cash_bal:.2f}",
        "avg_expense_last_7_days": f"{avg_exp7:.2f}",
        "warehouse_value_rub": warehouse_val,
        "cushion_days": cushion_days,
        "prev": {
            "income": f"{prev.income:.2f}",
            "expense": f"{prev.expense:.2f}",
            "profit": f"{prev.profit:.2f}",
            "avg_expense_per_day": f"{prev.avg_expense_per_day:.2f}",
            "avg_income_per_day": f"{prev.avg_income_per_day:.2f}",
        },
        "delta": {
            "income": _delta_pct(dash.income, prev.income),
            "expense": _delta_pct(dash.expense, prev.expense),
            "profit": _delta_pct(dash.profit, prev.profit),
            "avg_expense_per_day": _delta_pct(dash.avg_expense_per_day, prev.avg_expense_per_day),
            "avg_income_per_day": _delta_pct(dash.avg_income_per_day, prev.avg_income_per_day),
        },
        "by_day": dash.by_day,
        "expense_by_category": exp_cats,
        "income_by_category": inc_cats,
        "top_expense_categories": dash.top_expense_categories,
        "top_income_categories": dash.top_income_categories,
    }


# ── Finance settings ──────────────────────────────────────────────────────────────

@router.get("/api/finance/settings")
async def fin_settings_get(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    bal = await _get_cash_balance(session=session)
    return {"ok": True, "cash_balance": f"{bal:.2f}"}


@router.post("/api/finance/settings/cash_balance")
async def fin_settings_set_cash_balance(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    body = await request.json()
    raw = str(body.get("cash_balance") or "0").replace(" ", "").replace(",", ".")
    try:
        value = Decimal(raw)
    except Exception:
        raise HTTPException(status_code=400, detail={"ok": False, "error": "invalid_amount"})
    if value < 0:
        raise HTTPException(status_code=400, detail={"ok": False, "error": "invalid_amount"})
    await _set_cash_balance(session=session, value=value)
    await session.commit()
    return {"ok": True, "cash_balance": f"{value:.2f}"}


# ── Categories ────────────────────────────────────────────────────────────────

@router.get("/api/finance/categories")
async def fin_cats_list(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    type_filter = str(request.query_params.get("type") or "").strip() or None
    include_archived = str(request.query_params.get("include_archived") or "") == "1"
    cats = await _list_cats(session=session, type_filter=type_filter, include_archived=include_archived)
    return {"ok": True, "items": [
        {"id": int(c.id), "type": str(c.type), "name": str(c.name), "is_archived": bool(c.is_archived)}
        for c in cats
    ]}


@router.post("/api/finance/categories")
async def fin_cats_create(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    try:
        body = await request.json()
    except Exception:
        body = {}
    try:
        cat = await _create_cat(session=session, type=str(body.get("type") or ""), name=str(body.get("name") or ""))
    except ValueError as e:
        code = str(e)
        return JSONResponse({"ok": False, "error": code, "error_message": _human_error(code)}, status_code=400)
    return {"ok": True, "id": int(cat.id), "type": str(cat.type), "name": str(cat.name)}


@router.put("/api/finance/categories/{category_id}")
async def fin_cats_update(category_id: int, request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    try:
        body = await request.json()
    except Exception:
        body = {}
    try:
        cat = await _update_cat(session=session, category_id=category_id, name=body.get("name"), is_archived=body.get("is_archived"))
    except ValueError as e:
        code = str(e)
        if code == "not_found":
            raise HTTPException(status_code=404)
        return JSONResponse({"ok": False, "error": code, "error_message": _human_error(code)}, status_code=400)
    return {"ok": True, "id": int(cat.id), "name": str(cat.name), "is_archived": bool(cat.is_archived)}


# ── Operations ────────────────────────────────────────────────────────────────

def _parse_dt(s: str) -> datetime | None:
    if not s:
        return None
    s = s.strip()
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except Exception:
            return None


@router.get("/api/finance/operations")
async def fin_ops_list(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    qp = request.query_params
    try:
        limit = max(1, min(200, int(qp.get("limit") or 50)))
        offset = max(0, int(qp.get("offset") or 0))
        category_id = int(qp.get("category_id") or 0) or None
    except Exception:
        limit, offset, category_id = 50, 0, None
    items, total = await _list_ops(
        session=session,
        date_from=_parse_dt(str(qp.get("from") or "")),
        date_to=_parse_dt(str(qp.get("to") or "")),
        type_filter=str(qp.get("type") or "").strip() or None,
        category_id=category_id,
        search=str(qp.get("search") or "").strip() or None,
        limit=limit, offset=offset,
    )
    return {"ok": True, "items": items, "total": total, "limit": limit, "offset": offset}


@router.post("/api/finance/operations")
async def fin_ops_create(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    content_type = str(request.headers.get("content-type") or "")
    file_paths: list[str] = []
    if "multipart" in content_type:
        form = await request.form()
        body: dict = {k: v for k, v in form.items() if not hasattr(v, "filename")}
        for v in form.values():
            if hasattr(v, "filename") and v.filename:
                try:
                    import uuid, os
                    from pathlib import Path
                    data = await v.read()
                    ext = Path(str(v.filename)).suffix or ".jpg"
                    fname = f"finance_{uuid.uuid4().hex}{ext}"
                    dest = Path("/app/web/app/static/uploads/finance")
                    dest.mkdir(parents=True, exist_ok=True)
                    (dest / fname).write_bytes(data)
                    file_paths.append(f"/crm/static/uploads/finance/{fname}")
                except Exception:
                    pass
    else:
        try:
            body = await request.json()
        except Exception:
            body = {}
    try:
        amount = Decimal(str(body.get("amount") or "0").replace(",", "."))
        category_id = int(body.get("category_id") or 0) or None
        op = await _create_op(
            session=session,
            type=str(body.get("type") or ""),
            amount=amount,
            occurred_at=_parse_dt(str(body.get("occurred_at") or "")) or utc_now(),
            category_id=category_id,
            subcategory=str(body.get("subcategory") or "").strip() or None,
            counterparty=str(body.get("counterparty") or "").strip() or None,
            payment_method=str(body.get("payment_method") or "").strip() or None,
            comment=str(body.get("comment") or "").strip() or None,
            file_paths=file_paths,
        )
    except ValueError as e:
        code = str(e)
        return JSONResponse({"ok": False, "error": code, "error_message": _human_error(code)}, status_code=400)
    return {"ok": True, "id": int(op.id)}


@router.get("/api/finance/operations/{operation_id}")
async def fin_op_detail(operation_id: int, request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    op = await _get_op(session=session, operation_id=operation_id)
    if op is None:
        raise HTTPException(status_code=404)
    return {"ok": True, "operation": _serialize_operation(op)}


@router.put("/api/finance/operations/{operation_id}")
async def fin_op_update(operation_id: int, request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    try:
        body = await request.json()
    except Exception:
        body = {}
    try:
        amount_raw = body.get("amount")
        amount = Decimal(str(amount_raw).replace(",", ".")) if amount_raw is not None else None
        category_id = int(body.get("category_id") or 0) or None
        op = await _update_op(
            session=session, operation_id=operation_id,
            type=body.get("type"), amount=amount,
            occurred_at=_parse_dt(str(body.get("occurred_at") or "")),
            category_id=category_id, subcategory=body.get("subcategory"),
            counterparty=body.get("counterparty"),
            payment_method=body.get("payment_method"), comment=body.get("comment"),
        )
    except ValueError as e:
        code = str(e)
        if code == "not_found":
            raise HTTPException(status_code=404)
        return JSONResponse({"ok": False, "error": code, "error_message": _human_error(code)}, status_code=400)
    return {"ok": True, "id": int(op.id)}


@router.delete("/api/finance/operations/{operation_id}")
async def fin_op_delete(operation_id: int, request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    try:
        await _delete_op(session=session, operation_id=operation_id)
    except ValueError as e:
        code = str(e)
        if code == "not_found":
            raise HTTPException(status_code=404)
        return JSONResponse({"ok": False, "error": code, "error_message": _human_error(code)}, status_code=400)
    return {"ok": True}


# ── Export ────────────────────────────────────────────────────────────────────

_LABELS: dict[str, str] = {
    "id": "ID", "type": "Тип", "amount": "Сумма", "occurred_at": "Дата",
    "category_name": "Категория", "subcategory": "Подкатегория",
    "counterparty": "Контрагент", "payment_method": "Способ оплаты",
    "comment": "Комментарий", "actor_name": "Автор", "created_at": "Создано",
    "income": "Доходы", "expense": "Расходы", "profit": "Прибыль",
    "day": "День", "category": "Категория",
}


@router.post("/api/finance/export")
async def fin_export(request: Request, session: AsyncSession = Depends(get_db)):
    if not pin_valid(request):
        raise HTTPException(status_code=403)
    try:
        body = await request.json()
    except Exception:
        body = {}
    try:
        category_id = int(body.get("category_id") or 0) or None
    except Exception:
        category_id = None
    rows = await _export_ops(
        session=session,
        date_from=_parse_dt(str(body.get("from") or "")),
        date_to=_parse_dt(str(body.get("to") or "")),
        type_filter=str(body.get("type") or "").strip() or None,
        category_id=category_id,
        search=str(body.get("search") or "").strip() or None,
        fields=list(body.get("fields") or []) or None,
        aggregate_by=str(body.get("aggregate_by") or "").strip() or None,
    )
    fmt = str(body.get("format") or "csv").strip().lower()
    if fmt == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Финансы"
        if rows:
            headers = list(rows[0].keys())
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=_LABELS.get(h, h))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
            for ri, row in enumerate(rows, 2):
                for ci, h in enumerate(headers, 1):
                    v = row.get(h)
                    ws.cell(row=ri, column=ci, value=str(v) if v is not None else "")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=finance_export.xlsx"},
        )
    buf2 = io.StringIO()
    if rows:
        headers = list(rows[0].keys())
        writer = csv.DictWriter(buf2, fieldnames=headers, delimiter=";", extrasaction="ignore")
        writer.writerow({h: _LABELS.get(h, h) for h in headers})
        writer.writerows(rows)
    buf2.seek(0)
    return StreamingResponse(
        iter([buf2.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=finance_export.csv"},
    )
